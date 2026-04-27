"""
Daily price scraper for the FootJoy AU competitive price tracker.

Reads `Price Tracker Sheet.xlsx` from the repo root, scrapes each product URL
using retailer-specific parsers, and appends one row per product per retailer
per day to `data/prices.json`. Writes a run summary to `data/last_run.json`.

Run locally:
    pip install -r requirements.txt
    playwright install chromium
    python scripts/scrape_prices.py
"""

from __future__ import annotations

import json
import logging
import os
import random
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

# Optional: only imported if a page needs JS rendering
_playwright_available = True
try:
    from playwright.sync_api import sync_playwright
except ImportError:
    _playwright_available = False


# --- Config ----------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
SHEET_PATH = REPO_ROOT / "Price Tracker Sheet.xlsx"
DATA_DIR = REPO_ROOT / "data"
PRICES_PATH = DATA_DIR / "prices.json"
LAST_RUN_PATH = DATA_DIR / "last_run.json"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 30
PER_DOMAIN_DELAY_RANGE = (2.0, 3.5)  # seconds, randomized per request
FAILURE_THRESHOLD = 0.30  # exit nonzero if >30% of URLs fail

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("scraper")


# --- Data types ------------------------------------------------------------

@dataclass
class Product:
    category: str
    brand: str
    product: str
    retailer: str
    url: str


@dataclass
class PricePoint:
    date: str
    timestamp_utc: str
    category: str
    brand: str
    product: str
    retailer: str
    url: str
    price: Optional[float]
    price_was: Optional[float]
    currency: str
    on_sale: bool
    in_stock: bool
    status: str  # "ok" | "error" | "carry_forward"
    error: Optional[str]


# --- Spreadsheet reading ---------------------------------------------------

def load_products(sheet_path: Path) -> list[Product]:
    wb = openpyxl.load_workbook(sheet_path, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise RuntimeError(f"Expected 'Sheet1' tab in {sheet_path.name}")
    ws = wb["Sheet1"]

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    required = {"Category", "Brand", "Product Name", "Retailer", "URL"}
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"Sheet1 missing columns: {missing}")

    idx = {h: headers.index(h) for h in required}
    out: list[Product] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        category = row[idx["Category"]]
        url = row[idx["URL"]]
        if not category or not url:
            continue
        out.append(
            Product(
                category=str(category).strip(),
                brand=str(row[idx["Brand"]] or "").strip(),
                product=str(row[idx["Product Name"]] or "").strip(),
                retailer=str(row[idx["Retailer"]] or "").strip(),
                url=str(url).strip(),
            )
        )
    log.info("Loaded %d products from %s", len(out), sheet_path.name)
    return out


# --- Fetching --------------------------------------------------------------

_session = requests.Session()
_session.headers.update(
    {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-AU,en;q=0.9",
    }
)


def fetch_static(url: str) -> Optional[str]:
    try:
        r = _session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code >= 400:
            log.warning("HTTP %s for %s", r.status_code, url)
            return None
        return r.text
    except requests.RequestException as e:
        log.warning("requests error on %s: %s", url, e)
        return None


def fetch_rendered(url: str) -> Optional[str]:
    if not _playwright_available:
        log.warning("Playwright unavailable, skipping rendered fetch for %s", url)
        return None
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(user_agent=USER_AGENT, locale="en-AU")
            page = ctx.new_page()
            page.goto(url, timeout=REQUEST_TIMEOUT * 1000, wait_until="domcontentloaded")
            # Give SPA-style sites a moment to populate price nodes.
            page.wait_for_timeout(2500)
            html = page.content()
            browser.close()
            return html
    except Exception as e:
        log.warning("Playwright error on %s: %s", url, e)
        return None


# --- Parsing helpers -------------------------------------------------------

PRICE_NUM_RE = re.compile(r"[\d,]+\.\d{2}|[\d,]+")


def _to_float(text: str) -> Optional[float]:
    if not text:
        return None
    m = PRICE_NUM_RE.search(text.replace(",", ""))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_jsonld(html: str) -> Optional[dict]:
    """Find the first JSON-LD block that looks like a Product offer and return
    {price, currency, on_sale, in_stock, price_was}. Returns None if not found."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = tag.string or tag.get_text() or ""
        if not raw.strip():
            continue
        # Some sites emit invalid JSON (trailing commas, comments). Try strict first.
        candidates = []
        try:
            candidates.append(json.loads(raw))
        except json.JSONDecodeError:
            # Try to recover by extracting the first balanced object.
            try:
                cleaned = re.sub(r",\s*([}\]])", r"\1", raw)
                candidates.append(json.loads(cleaned))
            except json.JSONDecodeError:
                continue

        for blob in candidates:
            for obj in _walk_jsonld(blob):
                offer = _extract_offer(obj)
                if offer:
                    return offer
    return None


def _walk_jsonld(blob):
    if isinstance(blob, list):
        for item in blob:
            yield from _walk_jsonld(item)
    elif isinstance(blob, dict):
        yield blob
        for v in blob.values():
            if isinstance(v, (dict, list)):
                yield from _walk_jsonld(v)


def _extract_offer(obj: dict) -> Optional[dict]:
    """Return a normalized price dict if `obj` looks like a Product/Offer node."""
    t = obj.get("@type")
    types = t if isinstance(t, list) else [t]
    types = [str(x).lower() for x in types if x]

    offers = obj.get("offers")
    if "product" in types and offers:
        offer_node = offers[0] if isinstance(offers, list) else offers
        return _normalize_offer(offer_node)
    if any(x in ("offer", "aggregateoffer") for x in types):
        return _normalize_offer(obj)
    return None


def _normalize_offer(o) -> Optional[dict]:
    if not isinstance(o, dict):
        return None
    price = _to_float(str(o.get("price", "")))
    if price is None:
        # AggregateOffer style: lowPrice / highPrice
        price = _to_float(str(o.get("lowPrice", "")))
    currency = str(o.get("priceCurrency") or o.get("currency") or "AUD").upper()
    availability = str(o.get("availability", "")).lower()
    in_stock = "instock" in availability or availability == "" or "preorder" in availability
    if "outofstock" in availability or "soldout" in availability:
        in_stock = False
    price_was = _to_float(str(o.get("priceSpecification", {}).get("price", ""))) if isinstance(o.get("priceSpecification"), dict) else None
    on_sale = price_was is not None and price is not None and price_was > price
    if price is None:
        return None
    return {
        "price": price,
        "currency": currency,
        "in_stock": in_stock,
        "on_sale": on_sale,
        "price_was": price_was,
    }


# --- Retailer-specific parsers --------------------------------------------

def parse_golfbox(html: str) -> Optional[dict]:
    # Most golfbox.com.au products expose JSON-LD; some BigCommerce Stencil
    # product pages only expose Open Graph / schema.org meta tags (e.g.
    # FootJoy Pro/SL White/Blue/Red), so meta tags are checked second and the
    # noisy regex is the last resort.
    return (
        parse_jsonld(html)
        or _extract_meta_price(BeautifulSoup(html, "lxml"))
        or _golfbox_regex(html)
    )


def parse_oncourse(html: str) -> Optional[dict]:
    return (
        parse_jsonld(html)
        or _extract_meta_price(BeautifulSoup(html, "lxml"))
        or _golfbox_regex(html)
    )


def _golfbox_regex(html: str) -> Optional[dict]:
    # Fallback used by both JSON-LD retailers.
    m_price = re.search(r'"price"\s*:\s*"?([\d.]+)"?', html)
    m_curr = re.search(r'"priceCurrency"\s*:\s*"([A-Z]{3})"', html)
    if not m_price:
        return None
    return {
        "price": _to_float(m_price.group(1)),
        "currency": m_curr.group(1) if m_curr else "AUD",
        "in_stock": True,
        "on_sale": False,
        "price_was": None,
    }


def _find_balanced_object(text: str, start: int) -> Optional[str]:
    """Return the substring beginning at text[start] (which must be '{') and
    ending at the matching '}', respecting JSON string literals."""
    if start >= len(text) or text[start] != "{":
        return None
    depth = 0
    in_str = False
    escape = False
    for i in range(start, len(text)):
        c = text[i]
        if escape:
            escape = False
            continue
        if c == "\\":
            escape = True
            continue
        if c == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def _extract_datalayer_product(html: str) -> Optional[dict]:
    """FootJoy AU (Salesforce Commerce Cloud) embeds the price in a Google Tag
    Manager dataLayer.push call:

        dataLayer.push({"event":"productView","ecommerce":{"currencyCode":"AUD",
            "detail":{"products":[{"id":"017PSL","price":299, ...}]}}});

    Walk every dataLayer.push in the page, parse the object, and return the
    first productView with a positive price."""
    needle = "dataLayer.push("
    idx = 0
    while True:
        i = html.find(needle, idx)
        if i < 0:
            return None
        brace = html.find("{", i)
        if brace < 0:
            return None
        body = _find_balanced_object(html, brace)
        if not body:
            idx = brace + 1
            continue
        try:
            blob = json.loads(body)
        except json.JSONDecodeError:
            idx = brace + 1
            continue
        if isinstance(blob, dict) and blob.get("event") == "productView":
            ecom = blob.get("ecommerce") or {}
            detail = ecom.get("detail") or {}
            products = detail.get("products") or []
            if products and isinstance(products[0], dict):
                p = products[0]
                try:
                    price = float(p.get("price")) if p.get("price") is not None else None
                except (ValueError, TypeError):
                    price = None
                if price and price > 0:
                    return {
                        "price": price,
                        "currency": (ecom.get("currencyCode") or "AUD").upper(),
                        "in_stock": True,
                        "on_sale": False,
                        "price_was": None,
                    }
        idx = brace + 1


def _extract_meta_price(soup: BeautifulSoup) -> Optional[dict]:
    """Some themes expose the price via Open Graph or schema.org meta tags.

    Recognises the BigCommerce Stencil pattern used by golfbox.com.au, where
    the on-sale price lives in `product:price:amount` and the RRP lives in
    `og:price:standard_amount`."""
    meta = (
        soup.select_one('meta[itemprop="price"]')
        or soup.select_one('meta[property="product:price:amount"]')
        or soup.select_one('meta[property="og:price:amount"]')
    )
    if not meta or not meta.get("content"):
        return None
    price = _to_float(meta["content"])
    if price is None:
        return None
    curr_meta = (
        soup.select_one('meta[itemprop="priceCurrency"]')
        or soup.select_one('meta[property="product:price:currency"]')
        or soup.select_one('meta[property="og:price:currency"]')
    )
    currency = (curr_meta.get("content") if curr_meta and curr_meta.get("content") else "AUD").upper()

    # BigCommerce: og:price:standard_amount is the RRP; if it's higher, the
    # current `price` is a sale price.
    standard_meta = (
        soup.select_one('meta[property="og:price:standard_amount"]')
        or soup.select_one('meta[property="product:price:standard_amount"]')
    )
    standard = _to_float(standard_meta["content"]) if standard_meta and standard_meta.get("content") else None
    on_sale = standard is not None and standard > price

    return {
        "price": price,
        "currency": currency,
        "in_stock": True,
        "on_sale": on_sale,
        "price_was": standard if on_sale else None,
    }


def parse_footjoy(html: str) -> Optional[dict]:
    soup = BeautifulSoup(html, "lxml")

    # Stock detection (used by every successful path below).
    text_lc = soup.get_text(" ", strip=True).lower()
    in_stock = "out of stock" not in text_lc and "sold out" not in text_lc

    # 1. Original .price-sales selector (kept for back-compat with older themes).
    sales = soup.select_one(".product-price .price-sales, .price-sales")
    standard = soup.select_one(".product-price .price-standard, .price-standard")
    sales_val = _to_float(sales.get_text(" ", strip=True)) if sales else None
    standard_val = _to_float(standard.get_text(" ", strip=True)) if standard else None
    if sales_val is not None:
        on_sale = standard_val is not None and standard_val > sales_val
        return {
            "price": sales_val,
            "currency": "AUD",
            "in_stock": in_stock,
            "on_sale": on_sale,
            "price_was": standard_val if on_sale else None,
        }

    # 2. dataLayer.push productView (current footjoy.com.au markup).
    dl = _extract_datalayer_product(html)
    if dl:
        dl["in_stock"] = in_stock
        return dl

    # 3. Open Graph / schema.org meta tags.
    mt = _extract_meta_price(soup)
    if mt:
        mt["in_stock"] = in_stock
        return mt

    # 4. JSON-LD as a final fallback.
    ld = parse_jsonld(html)
    if ld:
        return ld

    return None


RETAILER_PARSERS = {
    "golf box": parse_golfbox,
    "on course": parse_oncourse,
    "footjoy": parse_footjoy,
}


def parser_for(retailer: str):
    key = retailer.strip().lower()
    return RETAILER_PARSERS.get(key)


# --- Data persistence ------------------------------------------------------

def load_existing_prices() -> list[dict]:
    if not PRICES_PATH.exists():
        return []
    try:
        return json.loads(PRICES_PATH.read_text())
    except json.JSONDecodeError:
        log.warning("Existing prices.json was unreadable, starting fresh")
        return []


def save_prices(rows: list[dict]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    rows_sorted = sorted(rows, key=lambda r: (r["date"], r["product"], r["retailer"]))
    PRICES_PATH.write_text(json.dumps(rows_sorted, indent=2, ensure_ascii=False))


def upsert(existing: list[dict], new: PricePoint) -> list[dict]:
    key = (new.date, new.product, new.retailer)
    out = [r for r in existing if (r["date"], r["product"], r["retailer"]) != key]
    out.append(asdict(new))
    return out


def last_known(existing: list[dict], product: str, retailer: str) -> Optional[dict]:
    matches = [
        r for r in existing
        if r["product"] == product and r["retailer"] == retailer and r.get("price") is not None
    ]
    if not matches:
        return None
    return max(matches, key=lambda r: r["date"])


# --- Main ------------------------------------------------------------------

def scrape_one(p: Product) -> tuple[Optional[dict], Optional[str]]:
    parser = parser_for(p.retailer)
    if not parser:
        return None, f"no parser registered for retailer '{p.retailer}'"

    html = fetch_static(p.url)
    if html:
        result = parser(html)
        if result and result.get("price") is not None:
            return result, None

    log.info("Static parse failed for %s, retrying with Playwright", p.url)
    html = fetch_rendered(p.url)
    if html:
        result = parser(html)
        if result and result.get("price") is not None:
            return result, None

    return None, "could not extract price from page"


def main() -> int:
    products = load_products(SHEET_PATH)
    existing = load_existing_prices()

    today = datetime.now(timezone.utc).date().isoformat()
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    successes: list[PricePoint] = []
    failures: list[dict] = []

    domain_last_hit: dict[str, float] = defaultdict(float)
    rng = random.Random()

    for p in products:
        domain = urlparse(p.url).netloc
        wait = PER_DOMAIN_DELAY_RANGE[0] + rng.random() * (
            PER_DOMAIN_DELAY_RANGE[1] - PER_DOMAIN_DELAY_RANGE[0]
        )
        elapsed = time.time() - domain_last_hit[domain]
        if elapsed < wait:
            time.sleep(wait - elapsed)
        domain_last_hit[domain] = time.time()

        log.info("[%s] %s — %s", p.retailer, p.product, p.url)
        result, err = scrape_one(p)

        if result and result.get("price") is not None:
            point = PricePoint(
                date=today,
                timestamp_utc=now_iso,
                category=p.category,
                brand=p.brand,
                product=p.product,
                retailer=p.retailer,
                url=p.url,
                price=result["price"],
                price_was=result.get("price_was"),
                currency=result.get("currency", "AUD"),
                on_sale=bool(result.get("on_sale")),
                in_stock=bool(result.get("in_stock", True)),
                status="ok",
                error=None,
            )
            successes.append(point)
            log.info("  -> %s %.2f%s", point.currency, point.price,
                     " (sale)" if point.on_sale else "")
        else:
            # Carry forward last known price so the line stays continuous.
            prev = last_known(existing, p.product, p.retailer)
            if prev:
                point = PricePoint(
                    date=today,
                    timestamp_utc=now_iso,
                    category=p.category,
                    brand=p.brand,
                    product=p.product,
                    retailer=p.retailer,
                    url=p.url,
                    price=prev["price"],
                    price_was=prev.get("price_was"),
                    currency=prev.get("currency", "AUD"),
                    on_sale=bool(prev.get("on_sale")),
                    in_stock=False,
                    status="carry_forward",
                    error=err,
                )
                successes.append(point)
                log.warning("  -> carry-forward last known price (%s)", err)
            else:
                # No prior data, log as failure.
                failures.append({"product": p.product, "retailer": p.retailer, "url": p.url, "error": err})
                log.warning("  -> FAILED (%s)", err)

    rows = existing
    for s in successes:
        rows = upsert(rows, s)
    save_prices(rows)

    total = len(products)
    fail_count = len(failures)
    success_count = total - fail_count
    last_run = {
        "timestamp_utc": now_iso,
        "total_urls": total,
        "successes": success_count,
        "failures": fail_count,
        "failure_details": failures,
    }
    LAST_RUN_PATH.write_text(json.dumps(last_run, indent=2))
    log.info("Run finished: %d ok, %d failed", success_count, fail_count)

    if total and (fail_count / total) > FAILURE_THRESHOLD:
        log.error("Failure rate %.0f%% exceeded threshold, exiting nonzero",
                  100 * fail_count / total)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
